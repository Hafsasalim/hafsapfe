"""Rapports BI — GET/POST /api/reports/"""
from fastapi import APIRouter, Depends, HTTPException
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from database import get_db
from models.models import Report, CoffeeSale, CoffeeType, PaymentMode, MLPrediction
from schemas.schemas import ReportCreate
from auth import get_current_user
from models.models import User
from datetime import datetime
import io

router = APIRouter()


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _apply_period_filter(q, period: str):
    """Filter a CoffeeSale-based query by period string.
    Formats: "1" (month number), "S1 2025", "S2 2025", "2025"
    """
    p = period.strip()
    parts = p.split()
    if parts[0] in ('S1', 'S2'):
        sem = parts[0]
        year = int(parts[1]) if len(parts) > 1 else None
        months = list(range(1, 7)) if sem == 'S1' else list(range(7, 13))
        q = q.filter(CoffeeSale.month_sort.in_(months))
        if year:
            q = q.filter(func.year(CoffeeSale.saleDate) == year)
    elif len(p) == 4 and p.isdigit():
        q = q.filter(func.year(CoffeeSale.saleDate) == int(p))
    else:
        month = parts[0]
        year = int(parts[1]) if len(parts) > 1 else None
        q = q.filter(CoffeeSale.month_sort == int(month))
        if year:
            q = q.filter(func.year(CoffeeSale.saleDate) == year)
    return q


def _compute_full_stats(period: str, db: Session):
    totals = _apply_period_filter(
        db.query(
            func.sum(CoffeeSale.amount).label('revenue'),
            func.count(CoffeeSale.saleId).label('count')
        ), period
    ).first()

    revenue = round(totals.revenue or 0, 2)
    count   = totals.count or 0

    top_prods = (
        _apply_period_filter(
            db.query(CoffeeType.name, CoffeeType.category,
                     func.sum(CoffeeSale.amount).label('revenue'),
                     func.count(CoffeeSale.saleId).label('count'))
            .join(CoffeeSale, CoffeeSale.coffeeId == CoffeeType.coffeeId),
            period
        )
        .group_by(CoffeeType.coffeeId, CoffeeType.name, CoffeeType.category)
        .order_by(func.sum(CoffeeSale.amount).desc())
        .limit(5).all()
    )

    cats = (
        _apply_period_filter(
            db.query(CoffeeType.category,
                     func.sum(CoffeeSale.amount).label('revenue'),
                     func.count(CoffeeSale.saleId).label('count'))
            .join(CoffeeSale, CoffeeSale.coffeeId == CoffeeType.coffeeId),
            period
        )
        .group_by(CoffeeType.category)
        .order_by(func.sum(CoffeeSale.amount).desc()).all()
    )
    total_rev = sum(r.revenue for r in cats) or 1

    pays = (
        _apply_period_filter(
            db.query(PaymentMode.type,
                     func.count(CoffeeSale.saleId).label('count'),
                     func.sum(CoffeeSale.amount).label('revenue'))
            .join(CoffeeSale, CoffeeSale.paymentId == PaymentMode.paymentId),
            period
        )
        .group_by(PaymentMode.type).all()
    )

    preds = (
        db.query(CoffeeType.name, MLPrediction.predictedPrice,
                 MLPrediction.confidence, MLPrediction.forecastDate)
        .join(CoffeeType, CoffeeType.coffeeId == MLPrediction.coffeeId)
        .order_by(MLPrediction.forecastDate).limit(10).all()
    )

    return {
        "period": period,
        "kpis": {
            "chiffre_affaires": revenue,
            "nombre_ventes":    count,
            "panier_moyen":     round(revenue / max(count, 1), 2),
        },
        "top_produits": [
            {"rang": i+1, "produit": r.name, "categorie": r.category,
             "ca": round(r.revenue, 2), "ventes": r.count}
            for i, r in enumerate(top_prods)
        ],
        "par_categorie": [
            {"categorie": r.category, "ca": round(r.revenue, 2),
             "ventes": r.count, "part_pct": round((r.revenue / total_rev)*100, 1)}
            for r in cats
        ],
        "par_paiement": [
            {"mode": r.type, "ventes": r.count, "ca": round(r.revenue, 2)}
            for r in pays
        ],
        "predictions": [
            {"cafe": r.name, "prix_predit": round(r.predictedPrice, 2),
             "confiance_pct": round((r.confidence or 0)*100, 1),
             "date": str(r.forecastDate)}
            for r in preds
        ],
    }


# ─────────────────────────────────────────────────────────────────────────────
# Excel builder
# ─────────────────────────────────────────────────────────────────────────────

def _build_excel(data: dict, created_at: str, notes: str) -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.chart import BarChart, PieChart, Reference
    from openpyxl.utils import get_column_letter

    # ── palette ──────────────────────────────────────────────────────────────
    C_BROWN = "C8913A"
    C_DARK  = "3D2D1E"
    C_WHITE = "FFFFFF"
    C_GREEN = "10B981"
    C_BLUE  = "3B82F6"
    C_PURP  = "8B5CF6"
    C_GREY  = "F5EFE6"
    C_ALT   = "FFFDF8"

    def fill(color):   return PatternFill("solid", fgColor=color)
    def bold(sz=10, color=C_DARK): return Font(bold=True, size=sz, color=color)
    def norm(sz=10, color=C_DARK): return Font(size=sz, color=color)
    def ctr():  return Alignment(horizontal='center', vertical='center', wrap_text=True)
    def lft():  return Alignment(horizontal='left',   vertical='center')

    def border_thin():
        s = Side(style='thin', color="DDD5CA")
        return Border(left=s, right=s, top=s, bottom=s)

    def write_header_row(ws, row, cols_labels, bg=C_DARK, col_start=1):
        for j, label in enumerate(cols_labels):
            c = ws.cell(row=row, column=col_start+j, value=label)
            c.font      = Font(bold=True, size=10, color=C_WHITE)
            c.fill      = fill(bg)
            c.alignment = ctr()
            c.border    = border_thin()

    def write_data_row(ws, row, values, col_start=1, fmt_map=None):
        for j, v in enumerate(values):
            c = ws.cell(row=row, column=col_start+j, value=v)
            c.fill      = fill(C_GREY if row % 2 == 0 else C_ALT)
            c.alignment = ctr()
            c.border    = border_thin()
            c.font      = norm(10)
            if fmt_map and j in fmt_map:
                c.number_format = fmt_map[j]

    def section_title(ws, row, text, col_start, col_end, bg=C_BROWN):
        ws.merge_cells(start_row=row, start_column=col_start,
                       end_row=row,   end_column=col_end)
        c = ws.cell(row=row, column=col_start, value=text)
        c.font      = Font(bold=True, size=12, color=C_WHITE)
        c.fill      = fill(bg)
        c.alignment = ctr()
        ws.row_dimensions[row].height = 26

    wb = Workbook()

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 1 — Résumé
    # ═══════════════════════════════════════════════════════════════════════════
    ws1 = wb.active
    ws1.title = "Résumé"
    ws1.sheet_view.showGridLines = False

    for col, w in [(1,3),(2,28),(3,22),(4,22),(5,22),(6,3)]:
        ws1.column_dimensions[get_column_letter(col)].width = w

    # Title
    ws1.row_dimensions[1].height = 8
    ws1.row_dimensions[2].height = 46
    ws1.merge_cells('B2:E2')
    c = ws1['B2']
    c.value     = f"☕  CoffeeBI — Rapport {data['period']}"
    c.font      = Font(bold=True, size=22, color=C_WHITE)
    c.fill      = fill(C_DARK)
    c.alignment = ctr()

    ws1.row_dimensions[3].height = 20
    ws1.merge_cells('B3:E3')
    c = ws1['B3']
    c.value     = f"Généré le {created_at[:16]}"
    c.font      = Font(size=9, color="888888", italic=True)
    c.alignment = ctr()

    ws1.row_dimensions[4].height = 14

    # KPI section title
    section_title(ws1, 5, "📊  INDICATEURS CLÉS", 2, 4, bg=C_BROWN)

    kpi_def = [
        ("Chiffre d'affaires", f"{data['kpis']['chiffre_affaires']:,.2f} MAD", C_GREEN),
        ("Nombre de ventes",   f"{data['kpis']['nombre_ventes']:,}",            C_BROWN),
        ("Panier moyen",       f"{data['kpis']['panier_moyen']:,.2f} MAD",      C_BLUE),
    ]
    ws1.row_dimensions[6].height = 22
    ws1.row_dimensions[7].height = 32
    ws1.row_dimensions[8].height = 14
    for i, (label, value, color) in enumerate(kpi_def):
        col = i + 2
        c = ws1.cell(row=6, column=col, value=label)
        c.font = Font(bold=True, size=9, color=C_WHITE)
        c.fill = fill(color)
        c.alignment = ctr()

        c = ws1.cell(row=7, column=col, value=value)
        c.font = Font(bold=True, size=14, color=color)
        c.fill = fill("FFFDF8")
        c.alignment = ctr()

    # Top products mini-table
    section_title(ws1, 9, "🏆  TOP 5 PRODUITS", 2, 5, bg=C_DARK)
    write_header_row(ws1, 10, ["Rang","Produit","CA (MAD)","Nb Ventes"], col_start=2)
    ws1.row_dimensions[10].height = 22

    for i, p in enumerate(data['top_produits']):
        row = 11 + i
        ws1.row_dimensions[row].height = 18
        write_data_row(ws1, row, [p['rang'], p['produit'], p['ca'], p['ventes']],
                       col_start=2, fmt_map={2: '#,##0.00'})

    # Bar chart — top products CA on résumé
    n_prod = len(data['top_produits'])
    if n_prod:
        bar = BarChart()
        bar.type   = "col"
        bar.title  = "Chiffre d'affaires par produit"
        bar.y_axis.title = "CA (MAD)"
        bar.style  = 10
        bar.width  = 18
        bar.height = 11
        # col 4 = CA (col_start=2, index 2 → absolute col 4)
        bar.add_data(Reference(ws1, min_col=4, min_row=10, max_row=10+n_prod),
                     titles_from_data=True)
        bar.set_categories(Reference(ws1, min_col=3, min_row=11, max_row=10+n_prod))
        bar.series[0].graphicalProperties.solidFill = C_BROWN
        ws1.add_chart(bar, f"B{12+n_prod}")

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 2 — Top Produits
    # ═══════════════════════════════════════════════════════════════════════════
    ws2 = wb.create_sheet("Top Produits")
    ws2.sheet_view.showGridLines = False
    for col, w in [(1,8),(2,26),(3,18),(4,18),(5,15)]:
        ws2.column_dimensions[get_column_letter(col)].width = w

    ws2.row_dimensions[1].height = 8
    ws2.row_dimensions[2].height = 40
    ws2.merge_cells('A2:E2')
    c = ws2['A2']
    c.value     = "🏆  Top 5 Produits"
    c.font      = Font(bold=True, size=18, color=C_WHITE)
    c.fill      = fill(C_BROWN)
    c.alignment = ctr()

    ws2.row_dimensions[3].height = 8
    ws2.row_dimensions[4].height = 24
    write_header_row(ws2, 4, ["Rang","Produit","Catégorie","CA (MAD)","Nb Ventes"])

    for i, p in enumerate(data['top_produits']):
        row = 5 + i
        ws2.row_dimensions[row].height = 20
        write_data_row(ws2, row,
                       [p['rang'], p['produit'], p['categorie'], p['ca'], p['ventes']],
                       fmt_map={3: '#,##0.00'})

    if n_prod:
        bar2 = BarChart()
        bar2.type  = "col"
        bar2.title = "Chiffre d'affaires par produit"
        bar2.y_axis.title = "CA (MAD)"
        bar2.style = 10; bar2.width = 20; bar2.height = 13
        bar2.add_data(Reference(ws2, min_col=4, min_row=4, max_row=4+n_prod),
                      titles_from_data=True)
        bar2.set_categories(Reference(ws2, min_col=2, min_row=5, max_row=4+n_prod))
        bar2.series[0].graphicalProperties.solidFill = C_BROWN
        ws2.add_chart(bar2, f"A{6+n_prod}")

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 3 — Catégories
    # ═══════════════════════════════════════════════════════════════════════════
    ws3 = wb.create_sheet("Par Catégorie")
    ws3.sheet_view.showGridLines = False
    for col, w in [(1,22),(2,18),(3,14),(4,12)]:
        ws3.column_dimensions[get_column_letter(col)].width = w

    ws3.row_dimensions[1].height = 8
    ws3.row_dimensions[2].height = 40
    ws3.merge_cells('A2:D2')
    c = ws3['A2']
    c.value     = "📂  Ventes par Catégorie"
    c.font      = Font(bold=True, size=18, color=C_WHITE)
    c.fill      = fill(C_BLUE)
    c.alignment = ctr()

    ws3.row_dimensions[3].height = 8
    ws3.row_dimensions[4].height = 24
    write_header_row(ws3, 4, ["Catégorie","CA (MAD)","Nb Ventes","Part (%)"])

    n_cat = len(data['par_categorie'])
    for i, cat in enumerate(data['par_categorie']):
        row = 5 + i
        ws3.row_dimensions[row].height = 20
        write_data_row(ws3, row,
                       [cat['categorie'], cat['ca'], cat['ventes'], cat['part_pct']],
                       fmt_map={1: '#,##0.00', 3: '0.0'})

    if n_cat:
        pie3 = PieChart()
        pie3.title = "Répartition du CA par catégorie"
        pie3.style = 10; pie3.width = 18; pie3.height = 13
        pie3.add_data(Reference(ws3, min_col=2, min_row=4, max_row=4+n_cat),
                      titles_from_data=True)
        pie3.set_categories(Reference(ws3, min_col=1, min_row=5, max_row=4+n_cat))
        ws3.add_chart(pie3, f"A{6+n_cat}")

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 4 — Mode de paiement
    # ═══════════════════════════════════════════════════════════════════════════
    ws4 = wb.create_sheet("Par Paiement")
    ws4.sheet_view.showGridLines = False
    for col, w in [(1,20),(2,18),(3,18)]:
        ws4.column_dimensions[get_column_letter(col)].width = w

    ws4.row_dimensions[1].height = 8
    ws4.row_dimensions[2].height = 40
    ws4.merge_cells('A2:C2')
    c = ws4['A2']
    c.value     = "💳  Ventes par Mode de Paiement"
    c.font      = Font(bold=True, size=18, color=C_WHITE)
    c.fill      = fill(C_GREEN)
    c.alignment = ctr()

    ws4.row_dimensions[3].height = 8
    ws4.row_dimensions[4].height = 24
    write_header_row(ws4, 4, ["Mode","Nb Ventes","CA (MAD)"])

    n_pay = len(data['par_paiement'])
    for i, pay in enumerate(data['par_paiement']):
        row = 5 + i
        ws4.row_dimensions[row].height = 20
        write_data_row(ws4, row, [pay['mode'], pay['ventes'], pay['ca']],
                       fmt_map={2: '#,##0.00'})

    if n_pay:
        pie4 = PieChart()
        pie4.title = "Répartition par mode de paiement"
        pie4.style = 10; pie4.width = 16; pie4.height = 12
        pie4.add_data(Reference(ws4, min_col=2, min_row=4, max_row=4+n_pay),
                      titles_from_data=True)
        pie4.set_categories(Reference(ws4, min_col=1, min_row=5, max_row=4+n_pay))
        ws4.add_chart(pie4, "E2")

        bar4 = BarChart()
        bar4.type  = "col"
        bar4.title = "CA par mode de paiement"
        bar4.y_axis.title = "CA (MAD)"
        bar4.style = 10; bar4.width = 16; bar4.height = 12
        bar4.add_data(Reference(ws4, min_col=3, min_row=4, max_row=4+n_pay),
                      titles_from_data=True)
        bar4.set_categories(Reference(ws4, min_col=1, min_row=5, max_row=4+n_pay))
        bar4.series[0].graphicalProperties.solidFill = C_GREEN
        ws4.add_chart(bar4, f"A{6+n_pay}")

    # ═══════════════════════════════════════════════════════════════════════════
    # SHEET 5 — Prédictions ML
    # ═══════════════════════════════════════════════════════════════════════════
    preds = data.get('predictions', [])
    if preds:
        ws5 = wb.create_sheet("Prédictions ML")
        ws5.sheet_view.showGridLines = False
        for col, w in [(1,22),(2,20),(3,14),(4,14)]:
            ws5.column_dimensions[get_column_letter(col)].width = w

        ws5.row_dimensions[1].height = 8
        ws5.row_dimensions[2].height = 40
        ws5.merge_cells('A2:D2')
        c = ws5['A2']
        c.value     = "🤖  Prédictions ML"
        c.font      = Font(bold=True, size=18, color=C_WHITE)
        c.fill      = fill(C_PURP)
        c.alignment = ctr()

        ws5.row_dimensions[3].height = 8
        ws5.row_dimensions[4].height = 24
        write_header_row(ws5, 4, ["Café","Prix Prédit (MAD)","Confiance (%)","Date"])

        n_pred = len(preds)
        for i, pred in enumerate(preds):
            row = 5 + i
            ws5.row_dimensions[row].height = 20
            write_data_row(ws5, row,
                           [pred['cafe'], pred['prix_predit'],
                            pred['confiance_pct'], pred['date']],
                           fmt_map={1: '#,##0.00', 2: '0.0'})

        bar5 = BarChart()
        bar5.type  = "col"
        bar5.title = "Prix prédits par café"
        bar5.y_axis.title = "Prix (MAD)"
        bar5.style = 10; bar5.width = 20; bar5.height = 13
        bar5.add_data(Reference(ws5, min_col=2, min_row=4, max_row=4+n_pred),
                      titles_from_data=True)
        bar5.set_categories(Reference(ws5, min_col=1, min_row=5, max_row=4+n_pred))
        bar5.series[0].graphicalProperties.solidFill = C_PURP
        ws5.add_chart(bar5, f"A{6+n_pred}")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ─────────────────────────────────────────────────────────────────────────────
# Routes
# ─────────────────────────────────────────────────────────────────────────────

@router.get("/")
def get_reports(db: Session = Depends(get_db), _: User = Depends(get_current_user)):
    reports = db.query(Report).order_by(Report.created_at.desc()).limit(20).all()
    return [
        {"id": r.reportId, "period": r.period, "total_sales": r.totalSales,
         "created_at": str(r.created_at), "notes": r.notes}
        for r in reports
    ]


@router.post("/generate")
def generate_report(
    body: ReportCreate,
    db: Session = Depends(get_db),
    current: User = Depends(get_current_user)
):
    stats = _compute_full_stats(body.period, db)
    total = stats["kpis"]["chiffre_affaires"]
    report = Report(
        period=body.period,
        totalSales=round(total, 2),
        generatedBy=current.userId,
        created_at=datetime.utcnow(),
        notes=body.notes or f"Rapport auto — {body.period}"
    )
    db.add(report)
    db.commit()
    db.refresh(report)
    return {"id": report.reportId, "period": report.period,
            "total_sales": report.totalSales, "created_at": str(report.created_at)}


@router.get("/full-data/{report_id}")
def get_full_report_data(
    report_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    report = db.query(Report).filter(Report.reportId == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport introuvable")
    data = _compute_full_stats(report.period, db)
    data["created_at"] = str(report.created_at)
    data["notes"]      = report.notes
    return data


@router.get("/export-excel/{report_id}")
def export_excel(
    report_id: int,
    db: Session = Depends(get_db),
    _: User = Depends(get_current_user)
):
    report = db.query(Report).filter(Report.reportId == report_id).first()
    if not report:
        raise HTTPException(status_code=404, detail="Rapport introuvable")

    data = _compute_full_stats(report.period, db)
    excel_bytes = _build_excel(data, str(report.created_at), report.notes or "")

    filename = f"rapport_{report.period.replace(' ', '_')}_coffeebi.xlsx"
    return StreamingResponse(
        excel_bytes,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
